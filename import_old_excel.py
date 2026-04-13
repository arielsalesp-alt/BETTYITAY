from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


SOURCE = Path(r"C:\Users\user\OneDrive\שולחן העבודה\מחשב\אישי\התחשבנות חודשית בטי ואיתי.xlsx")
DEST = Path(__file__).with_name("archive-data.js")

SKIP_NAMES = {"סהכ", 'סה"כ', "בטי", "איתי", "0.5", "0.6", "0.4", "הפרש", "חודש", "סכום"}


def clean(value: Any) -> str:
    return str(value or "").strip()


def number(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def category_name(ws: Any, row: int, col: int, fallback: str) -> str:
    for lookup_row in range(row - 1, 0, -1):
        value = clean(ws.cell(lookup_row, col).value)
        if value and "שם ההוצאה" not in value:
            return value
        if lookup_row <= row - 2:
            break
    return fallback


def parse_expense_sheet(ws: Any) -> dict[str, Any]:
    categories = []
    seen_headers = set()

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column - 1):
            header = clean(ws.cell(row, col).value)
            betty_header = clean(ws.cell(row, col + 1).value)
            itai_header = clean(ws.cell(row, col + 2).value)
            if "שם ההוצאה" not in header or "בטי" not in betty_header or "איתי" not in itai_header:
                continue
            if (row, col) in seen_headers:
                continue
            seen_headers.add((row, col))

            items = []
            for item_row in range(row + 1, ws.max_row + 1):
                name = clean(ws.cell(item_row, col).value)
                betty = number(ws.cell(item_row, col + 1).value)
                itai = number(ws.cell(item_row, col + 2).value)

                if name and "שם ההוצאה" in name:
                    break
                if name in SKIP_NAMES or name.startswith("="):
                    continue
                if not name and betty == 0 and itai == 0:
                    continue

                items.append({"name": name, "betty": betty, "itai": itai})

            if items:
                categories.append(
                    {
                        "key": f"archive-{row}-{col}",
                        "name": category_name(ws, row, col, "הוצאות"),
                        "items": items,
                    }
                )

    if not categories:
        rows = []
        for row in range(2, ws.max_row + 1):
            name = clean(ws.cell(row, 1).value)
            betty = number(ws.cell(row, 2).value)
            itai = number(ws.cell(row, 3).value)
            if name and name not in SKIP_NAMES and (betty or itai):
                rows.append({"name": name, "betty": betty, "itai": itai})
        if rows:
            categories.append({"key": "archive-main", "name": "הוצאות", "items": rows})

    return {"name": ws.title, "income": {"itai": 0, "betty": 0}, "categories": categories, "source": "ארכיון"}


def json_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean(value)


def parse_debts(ws: Any) -> dict[str, Any]:
    bankruptcy = []
    for row in range(3, ws.max_row + 1):
        name = clean(ws.cell(row, 1).value)
        amount = number(ws.cell(row, 2).value)
        if "סה" in name:
            continue
        if name and amount:
            bankruptcy.append({"name": name, "amount": amount})

    years = []
    for year, month_col, amount_col in [(2022, 4, 5), (2023, 7, 8), (2024, 10, 11), (2025, 13, 14)]:
        entries = []
        for row in range(3, ws.max_row + 1):
            raw_month = clean(ws.cell(row, month_col).value)
            if "סה" in raw_month or "ñä" in raw_month.lower():
                continue
            month = json_date(ws.cell(row, month_col).value)
            amount = number(ws.cell(row, amount_col).value)
            if month or amount:
                entries.append({"month": month, "amount": amount})
        years.append({"year": year, "entries": entries})

    return {"name": ws.title, "bankruptcy": bankruptcy, "years": years}


def main() -> None:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    months = []
    debts = None

    for ws in workbook.worksheets:
        if "חובות" in ws.title:
            debts = parse_debts(ws)
            continue
        month = parse_expense_sheet(ws)
        if month["categories"]:
            months.append(month)

    payload = {
        "sourceFile": str(SOURCE),
        "generatedFrom": SOURCE.name,
        "months": months,
        "debts": debts,
    }
    DEST.write_text(
        "window.EXPENSE_ARCHIVE_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"wrote {DEST.resolve()} with {len(months)} archive months")


if __name__ == "__main__":
    main()
